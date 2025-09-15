#!/usr/bin/env python3
"""
Excel文件处理示例
演示如何使用contract_splitter处理Excel文件，特别是法律文档
"""

import sys
import os
from pathlib import Path

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

def example_excel_processor():
    """示例：Excel处理器基本功能"""
    print("📊 Excel处理器基本功能示例")
    print("=" * 50)
    
    from contract_splitter.excel_processor import ExcelProcessor
    
    processor = ExcelProcessor()
    
    print(f"可用的Excel处理库: {processor.available_libraries}")
    
    # 测试文件格式检测
    test_files = [
        "document.xlsx",
        "spreadsheet.xls", 
        "workbook.xlsm",
        "template.xltx",
        "report.pdf"
    ]
    
    print(f"\n文件格式检测:")
    for file_path in test_files:
        is_excel = processor.is_excel_file(file_path)
        print(f"  {file_path}: {'✅ Excel' if is_excel else '❌ 非Excel'}")


def example_simple_chunking():
    """示例：使用SimpleChunker处理Excel文件"""
    print("\n🔧 SimpleChunker Excel处理示例")
    print("=" * 50)
    
    from contract_splitter import simple_chunk_file, simple_chunk_text
    
    # 模拟Excel提取的文本内容
    excel_content = """
【工作表: 证券公司分类评价指标】
★ 指标类别 | 指标名称 | 评分标准 | 权重
★ 风险管理能力 | 净资本充足率 | 优秀≥150%，良好120%-150%，一般100%-120%，较差<100% | 30%
★ 风险管理能力 | 流动性覆盖率 | 优秀≥120%，良好100%-120%，一般80%-100%，较差<80% | 20%
★ 持续合规状况 | 合规检查结果 | 无重大问题为优秀，有一般问题为良好，有较严重问题为一般，有重大问题为较差 | 25%
★ 市场竞争力 | 市场份额 | 行业前10%为优秀，前20%为良好，前50%为一般，其他为较差 | 25%

【工作表: 分类结果应用】
第一条 根据评价结果，证券公司分为A、B、C、D四类。
第二条 A类证券公司在业务准入、产品创新等方面享受优先政策。
第三条 B类证券公司按照常规监管要求执行。
第四条 C类证券公司需要加强风险管控，限制部分业务。
第五条 D类证券公司需要重点监管，严格限制业务范围。
    """
    
    print(f"原始内容长度: {len(excel_content)} 字符")
    
    # 使用不同配置进行chunking
    configs = [
        {"max_chunk_size": 200, "overlap_ratio": 0.1, "name": "小chunks"},
        {"max_chunk_size": 400, "overlap_ratio": 0.15, "name": "中chunks"},
        {"max_chunk_size": 600, "overlap_ratio": 0.2, "name": "大chunks"},
    ]
    
    for config in configs:
        print(f"\n🔧 {config['name']} (max_size={config['max_chunk_size']}, overlap={config['overlap_ratio']})")
        
        chunks = simple_chunk_text(
            excel_content,
            max_chunk_size=config['max_chunk_size'],
            overlap_ratio=config['overlap_ratio']
        )
        
        print(f"生成 {len(chunks)} 个chunks")
        
        for i, chunk in enumerate(chunks[:2]):
            content = chunk['content'].strip()
            print(f"  Chunk {i+1} ({len(content)}字符): {content[:80]}...")


def example_excel_splitter():
    """示例：使用ExcelSplitter处理Excel文件"""
    print("\n📋 ExcelSplitter处理示例")
    print("=" * 50)
    
    from contract_splitter import ExcelSplitter
    
    # 创建测试Excel文件
    test_file = create_sample_excel_file()
    
    if not test_file:
        print("⚠️ 无法创建测试Excel文件，跳过此示例")
        return
    
    print(f"使用测试文件: {test_file}")
    
    # 测试不同的提取模式
    extract_modes = [
        ("legal_content", "法律内容模式"),
        ("table_structure", "表格结构模式"),
        ("all_content", "全部内容模式")
    ]
    
    for mode, description in extract_modes:
        print(f"\n🔍 {description} ({mode})")
        
        splitter = ExcelSplitter(
            max_tokens=300,
            extract_mode=mode
        )
        
        try:
            sections = splitter.split(test_file)
            
            print(f"生成 {len(sections)} 个sections")
            
            for i, section in enumerate(sections[:2]):
                print(f"  Section {i+1}:")
                print(f"    标题: {section['heading']}")
                print(f"    级别: {section['level']}")
                print(f"    类型: {section.get('section_type', 'unknown')}")
                print(f"    内容长度: {len(section['content'])} 字符")
                print(f"    内容预览: {section['content'][:60]}...")
                
        except Exception as e:
            print(f"  ❌ 处理失败: {e}")


def example_factory_usage():
    """示例：使用SplitterFactory处理Excel文件"""
    print("\n🏭 SplitterFactory Excel处理示例")
    print("=" * 50)
    
    from contract_splitter import SplitterFactory
    
    factory = SplitterFactory()
    
    # 显示支持的格式
    supported_formats = factory.get_supported_formats()
    excel_formats = [fmt for fmt in supported_formats if fmt.startswith('xl')]
    
    print(f"支持的Excel格式: {excel_formats}")
    
    # 创建测试文件
    test_file = create_sample_excel_file()
    
    if test_file:
        print(f"\n使用Factory处理: {test_file}")
        
        try:
            # 自动检测格式并创建splitter
            splitter = factory.create_splitter(test_file, max_tokens=400)
            
            print(f"自动创建的splitter类型: {type(splitter).__name__}")
            
            # 分割文档
            sections = splitter.split(test_file)
            
            print(f"生成 {len(sections)} 个sections")
            
            # 展平为chunks
            chunks = splitter.flatten(sections)
            
            print(f"展平为 {len(chunks)} 个chunks")
            
            for i, chunk in enumerate(chunks[:2]):
                print(f"  Chunk {i+1} ({len(chunk)}字符): {chunk[:60]}...")
                
        except Exception as e:
            print(f"❌ Factory处理失败: {e}")


def create_sample_excel_file():
    """创建示例Excel文件"""
    try:
        import openpyxl
    except ImportError:
        print("⚠️ openpyxl未安装，无法创建Excel文件")
        return None
    
    from openpyxl import Workbook
    
    # 创建工作簿
    wb = Workbook()
    
    # 第一个工作表：法律条文
    ws1 = wb.active
    ws1.title = "法律条文"
    
    legal_data = [
        ["条文编号", "条文内容", "适用范围", "生效日期"],
        ["第一条", "为了规范证券公司分类监管，合理配置监管资源，提高监管效率，促进证券业健康发展，制定本规定。", "全部证券公司", "2023-01-01"],
        ["第二条", "本规定适用于在中华人民共和国境内依法设立的证券公司。", "境内证券公司", "2023-01-01"],
        ["第三条", "中国证监会及其派出机构依照本规定对证券公司进行分类监管。", "监管机构", "2023-01-01"],
        ["第四条", "证券公司分类监管是指中国证监会根据证券公司风险管理能力、持续合规状况等因素，将证券公司分为不同类别。", "分类标准", "2023-01-01"],
        ["第五条", "证券公司分类评价每年进行一次，评价基准日为每年的12月31日。", "评价周期", "2023-01-01"],
    ]
    
    for row_data in legal_data:
        ws1.append(row_data)
    
    # 第二个工作表：评价指标
    ws2 = wb.create_sheet("评价指标")
    
    indicator_data = [
        ["指标类别", "指标名称", "评分标准", "权重", "备注"],
        ["风险管理能力", "净资本充足率", "优秀≥150%，良好120%-150%，一般100%-120%，较差<100%", "30%", "核心指标"],
        ["风险管理能力", "流动性覆盖率", "优秀≥120%，良好100%-120%，一般80%-100%，较差<80%", "20%", "重要指标"],
        ["持续合规状况", "合规检查结果", "无重大问题为优秀，有一般问题为良好，有较严重问题为一般，有重大问题为较差", "25%", "关键指标"],
        ["市场竞争力", "市场份额", "行业前10%为优秀，前20%为良好，前50%为一般，其他为较差", "15%", "参考指标"],
        ["市场竞争力", "客户满意度", "≥90%为优秀，80%-90%为良好，70%-80%为一般，<70%为较差", "10%", "辅助指标"],
    ]
    
    for row_data in indicator_data:
        ws2.append(row_data)
    
    # 第三个工作表：分类结果
    ws3 = wb.create_sheet("分类结果")
    
    result_data = [
        ["分类", "条件", "监管措施", "业务权限"],
        ["A类", "评分≥80分", "常规监管", "全部业务"],
        ["B类", "评分60-80分", "重点关注", "大部分业务"],
        ["C类", "评分40-60分", "加强监管", "限制部分业务"],
        ["D类", "评分<40分", "重点监管", "严格限制业务"],
    ]
    
    for row_data in result_data:
        ws3.append(row_data)
    
    # 保存文件
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    
    test_file = output_dir / "sample_legal_excel.xlsx"
    wb.save(test_file)
    
    print(f"✅ 示例Excel文件已创建: {test_file}")
    return str(test_file)


def example_real_world_usage():
    """示例：实际应用场景"""
    print("\n🌍 实际应用场景示例")
    print("=" * 50)
    
    print("Excel文件处理的典型应用场景:")
    print("1. 📋 法律法规条文表格 - 提取条文内容进行分析")
    print("2. 📊 监管指标数据 - 解析评价标准和权重")
    print("3. 📈 合规检查结果 - 提取检查项目和结果")
    print("4. 📑 政策文件附表 - 处理政策配套的数据表格")
    print("5. 🔍 案例分析数据 - 提取案例信息进行学习")
    
    print(f"\n推荐配置:")
    print("- 法律条文表格: ExcelSplitter(extract_mode='legal_content', max_tokens=1500)")
    print("- 数据表格: ExcelSplitter(extract_mode='table_structure', max_tokens=1000)")
    print("- 混合内容: ExcelSplitter(extract_mode='all_content', max_tokens=2000)")
    
    print(f"\n最佳实践:")
    print("1. 🔍 先检查Excel文件结构，选择合适的提取模式")
    print("2. 📏 根据内容复杂度调整max_tokens参数")
    print("3. 🔄 对于大型表格，考虑增加overlap提高连续性")
    print("4. ✅ 处理后验证内容完整性和结构正确性")


def main():
    """主函数"""
    print("🚀 Excel文件处理示例")
    print("=" * 80)
    
    # 运行各种示例
    example_excel_processor()
    example_simple_chunking()
    example_excel_splitter()
    example_factory_usage()
    example_real_world_usage()
    
    print(f"\n🎯 总结")
    print("=" * 80)
    print("Excel文件处理功能已成功集成到contract_splitter包中！")
    print("主要特性:")
    print("✅ 支持多种Excel格式 (.xlsx, .xls, .xlsm等)")
    print("✅ 智能法律内容识别和提取")
    print("✅ 灵活的表格结构处理")
    print("✅ 与现有chunking系统无缝集成")
    print("✅ 支持多种提取模式和配置选项")
    
    print(f"\n💡 使用建议:")
    print("- 对于法律文档Excel，推荐使用legal_content模式")
    print("- 对于数据表格，推荐使用table_structure模式")
    print("- 可以通过SimpleChunker快速处理，或使用ExcelSplitter获得更多控制")


if __name__ == "__main__":
    main()
