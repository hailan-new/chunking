#!/usr/bin/env python3
"""
测试法律条文提取功能
专门测试新的law_articles提取模式，处理法规名称-条文-内容格式
"""

import os
import sys
import tempfile
import logging
from pathlib import Path

# 添加项目根目录到路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from contract_splitter.excel_processor import ExcelProcessor
from contract_splitter.excel_splitter import ExcelSplitter

# 设置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def create_test_excel_file():
    """创建测试用的Excel文件，模拟法规名称-条文-内容格式"""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "法律条文"
        
        # 添加标题行
        ws['A1'] = "法规名称"
        ws['B1'] = "条文内容"

        # 添加测试数据（2列格式：法规名称 + 条文内容）
        test_data = [
            ["中华人民共和国工业产品生产许可证管理条例", "第一条 为了保证直接关系公共安全、人体健康、生命财产安全的重要工业产品的质量安全，贯彻国家产业政策，促进经济社会发展，制定本条例。"],
            ["中华人民共和国工业产品生产许可证管理条例", "第二条 国家对直接关系公共安全、人体健康、生命财产安全的重要工业产品，实行生产许可证制度。"],
            ["中华人民共和国工业产品生产许可证管理条例", "第三条 企业未依照本条例规定取得生产许可证的，不得生产列入目录的产品。任何单位和个人不得销售或者在经营活动中使用未取得生产许可证的列入目录的产品。"],
            ["中华人民共和国工业产品生产许可证管理条例", "第四条 国务院质量技术监督部门负责全国工业产品生产许可证统一管理工作。国务院有关部门在各自的职责范围内负责相关工业产品生产许可证管理工作。"],
            ["中华人民共和国工业产品生产许可证管理条例", "第五条 县级以上地方质量技术监督部门负责本行政区域内工业产品生产许可证管理工作。"],
        ]

        # 写入数据
        for i, (law_name, article_content) in enumerate(test_data, start=2):
            ws[f'A{i}'] = law_name
            ws[f'B{i}'] = article_content
        
        # 保存到临时文件
        temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(temp_file.name)
        wb.close()
        
        return temp_file.name
        
    except ImportError:
        logger.error("openpyxl未安装，无法创建测试Excel文件")
        return None


def test_excel_processor_law_articles():
    """测试ExcelProcessor的law_articles模式"""
    logger.info("=== 测试ExcelProcessor的law_articles模式 ===")
    
    # 创建测试文件
    test_file = create_test_excel_file()
    if not test_file:
        logger.error("无法创建测试文件")
        return False
    
    try:
        # 初始化处理器
        processor = ExcelProcessor()
        
        # 测试law_articles模式
        logger.info("测试law_articles提取模式...")
        text_content = processor.extract_text(test_file, extract_mode="law_articles")
        
        if text_content:
            logger.info(f"提取的文本长度: {len(text_content)} 字符")
            logger.info("提取的内容预览:")
            print("=" * 60)
            print(text_content)
            print("=" * 60)
            
            # 验证内容格式（特殊格式：法规名称单独一块，条文分别成块）
            expected_patterns = [
                "【LAW_NAME】",
                "中华人民共和国工业产品生产许可证管理条例",
                "【ARTICLE】第一条",
                "【ARTICLE】第二条",
                "为了保证直接关系公共安全",
                "国家对直接关系公共安全"
            ]
            
            success = True
            for pattern in expected_patterns:
                if pattern not in text_content:
                    logger.error(f"未找到期望的内容: {pattern}")
                    success = False
                else:
                    logger.info(f"✓ 找到期望的内容: {pattern}")
            
            return success
        else:
            logger.error("提取的文本内容为空")
            return False
            
    except Exception as e:
        logger.error(f"测试过程中出现错误: {e}")
        return False
    finally:
        # 清理临时文件
        if os.path.exists(test_file):
            os.unlink(test_file)


def test_excel_splitter_law_articles():
    """测试ExcelSplitter的law_articles模式"""
    logger.info("=== 测试ExcelSplitter的law_articles模式 ===")
    
    # 创建测试文件
    test_file = create_test_excel_file()
    if not test_file:
        logger.error("无法创建测试文件")
        return False
    
    try:
        # 初始化分割器
        splitter = ExcelSplitter(
            extract_mode="law_articles",
            max_tokens=1000,
            overlap=50
        )
        
        # 执行分割
        logger.info("执行法律条文分割...")
        chunks = splitter.split(test_file)
        
        if chunks:
            logger.info(f"生成的块数量: {len(chunks)}")
            
            # 验证每个块
            for i, chunk in enumerate(chunks):
                logger.info(f"\n--- 块 {i+1} ---")
                logger.info(f"标题: {chunk.get('heading', 'N/A')}")
                logger.info(f"内容长度: {len(chunk.get('content', ''))}")
                logger.info(f"类型: {chunk.get('section_type', 'N/A')}")
                logger.info(f"内容预览: {chunk.get('content', '')[:100]}...")
            
            # 验证是否每个条文都成为了独立的块
            law_article_chunks = [chunk for chunk in chunks if chunk.get('section_type') == 'law_article']
            logger.info(f"法律条文块数量: {len(law_article_chunks)}")
            
            # 应该有1个法规名称块 + 5个法律条文块（第一条到第五条）
            law_name_chunks = [chunk for chunk in chunks if chunk.get('section_type') == 'law_name']

            if len(law_name_chunks) >= 1 and len(law_article_chunks) >= 5:
                logger.info("✓ 成功生成了期望数量的法律条文块")

                # 验证法规名称块
                law_name_chunk = law_name_chunks[0]
                if "中华人民共和国工业产品生产许可证管理条例" in law_name_chunk.get('content', ''):
                    logger.info("✓ 法规名称块包含正确的法规名称")
                else:
                    logger.error("✗ 法规名称块不包含正确的法规名称")
                    return False

                # 验证第一个条文块
                first_article_chunk = law_article_chunks[0]
                if "第一条" in first_article_chunk.get('heading', ''):
                    logger.info("✓ 第一个条文块标题包含条文号")
                else:
                    logger.error("✗ 第一个条文块标题不包含条文号")
                    return False

                return True
            else:
                logger.error(f"期望1个法规名称块和5个法律条文块，实际得到{len(law_name_chunks)}个法规名称块和{len(law_article_chunks)}个法律条文块")
                return False
        else:
            logger.error("未生成任何块")
            return False
            
    except Exception as e:
        logger.error(f"测试过程中出现错误: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        # 清理临时文件
        if os.path.exists(test_file):
            os.unlink(test_file)


def main():
    """主测试函数"""
    logger.info("开始测试法律条文提取功能")
    
    # 测试ExcelProcessor
    processor_success = test_excel_processor_law_articles()
    
    # 测试ExcelSplitter
    splitter_success = test_excel_splitter_law_articles()
    
    # 总结结果
    logger.info("\n" + "=" * 60)
    logger.info("测试结果总结:")
    logger.info(f"ExcelProcessor测试: {'✓ 通过' if processor_success else '✗ 失败'}")
    logger.info(f"ExcelSplitter测试: {'✓ 通过' if splitter_success else '✗ 失败'}")
    
    if processor_success and splitter_success:
        logger.info("🎉 所有测试通过！法律条文提取功能工作正常。")
        return True
    else:
        logger.error("❌ 部分测试失败，请检查实现。")
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
