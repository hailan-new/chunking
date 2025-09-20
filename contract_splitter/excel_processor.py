#!/usr/bin/env python3
"""
Excel文档处理器
专门处理Excel文件(.xlsx, .xls, .xlsm)，提取结构化内容并针对法律文档进行优化
"""

import os
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
import tempfile

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """
    Excel文档处理器
    支持多种Excel格式，提供文本提取和结构化处理功能
    """
    
    def __init__(self):
        """初始化Excel处理器"""
        self.available_libraries = self._check_available_libraries()
        logger.info(f"Excel处理器初始化完成，可用库: {list(self.available_libraries.keys())}")
    
    def _check_available_libraries(self) -> Dict[str, bool]:
        """检查可用的Excel处理库"""
        libraries = {
            'openpyxl': False,
            'xlrd': False,
            'pandas': False,
            'xlwings': False
        }
        
        # 检查openpyxl (推荐用于.xlsx)
        try:
            import openpyxl
            libraries['openpyxl'] = True
            logger.info("✓ openpyxl可用 - 支持.xlsx/.xlsm文件")
        except ImportError:
            logger.warning("✗ openpyxl不可用")
        
        # 检查xlrd (用于.xls)
        try:
            import xlrd
            libraries['xlrd'] = True
            logger.info("✓ xlrd可用 - 支持.xls文件")
        except ImportError:
            logger.warning("✗ xlrd不可用")
        
        # 检查pandas (通用处理)
        try:
            import pandas as pd
            libraries['pandas'] = True
            logger.info("✓ pandas可用 - 通用Excel处理")
        except ImportError:
            logger.warning("✗ pandas不可用")
        
        # 检查xlwings (Windows Excel COM)
        try:
            import xlwings
            libraries['xlwings'] = True
            logger.info("✓ xlwings可用 - Windows Excel COM支持")
        except ImportError:
            pass  # xlwings是可选的
        
        return libraries
    
    def is_excel_file(self, file_path: str) -> bool:
        """
        检测文件是否为Excel格式
        
        Args:
            file_path: 文件路径
            
        Returns:
            True表示是Excel文件
        """
        if not os.path.exists(file_path):
            return False
        
        # 检查文件扩展名
        ext = Path(file_path).suffix.lower()
        excel_extensions = ['.xlsx', '.xls', '.xlsm', '.xltx', '.xltm']
        
        if ext in excel_extensions:
            return True
        
        # 检查文件头（ZIP格式检测）
        try:
            with open(file_path, 'rb') as f:
                header = f.read(4)
                # Excel 2007+ 文件是ZIP格式
                if header == b'PK\x03\x04':
                    return True
                # Excel 97-2003 文件头
                elif header[:2] == b'\xd0\xcf':
                    return True
        except Exception:
            pass
        
        return False
    
    def extract_text(self, file_path: str, extract_mode: str = "legal_content") -> Optional[str]:
        """
        从Excel文件提取文本内容
        
        Args:
            file_path: Excel文件路径
            extract_mode: 提取模式 ("legal_content", "law_articles", "table_structure", "all_content")
                - "legal_content": 提取法律相关内容
                - "law_articles": 专门处理法规名称-条文-内容格式（A列法规名称，B列条文号，C列内容）
                - "table_structure": 提取表格结构
                - "all_content": 提取所有内容
            
        Returns:
            提取的文本内容，失败返回None
        """
        if not self.is_excel_file(file_path):
            logger.error(f"不是有效的Excel文件: {file_path}")
            return None
        
        file_ext = Path(file_path).suffix.lower()
        
        # 根据文件类型选择处理方法
        if file_ext in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
            return self._extract_from_xlsx(file_path, extract_mode)
        elif file_ext == '.xls':
            return self._extract_from_xls(file_path, extract_mode)
        else:
            logger.warning(f"不支持的Excel格式: {file_ext}")
            return None
    
    def _extract_from_xlsx(self, file_path: str, extract_mode: str) -> Optional[str]:
        """从.xlsx/.xlsm文件提取文本"""
        # 优先使用openpyxl
        if self.available_libraries.get('openpyxl'):
            try:
                return self._extract_with_openpyxl(file_path, extract_mode)
            except Exception as e:
                logger.warning(f"openpyxl提取失败: {e}")
        
        # 备选：使用pandas
        if self.available_libraries.get('pandas'):
            try:
                return self._extract_with_pandas(file_path, extract_mode)
            except Exception as e:
                logger.warning(f"pandas提取失败: {e}")
        
        logger.error(f"无法处理Excel文件: {file_path}")
        return None
    
    def _extract_from_xls(self, file_path: str, extract_mode: str) -> Optional[str]:
        """从.xls文件提取文本"""
        # 优先使用xlrd
        if self.available_libraries.get('xlrd'):
            try:
                return self._extract_with_xlrd(file_path, extract_mode)
            except Exception as e:
                logger.warning(f"xlrd提取失败: {e}")
        
        # 备选：使用pandas
        if self.available_libraries.get('pandas'):
            try:
                return self._extract_with_pandas(file_path, extract_mode)
            except Exception as e:
                logger.warning(f"pandas提取失败: {e}")
        
        logger.error(f"无法处理Excel文件: {file_path}")
        return None
    
    def _extract_with_openpyxl(self, file_path: str, extract_mode: str) -> str:
        """使用openpyxl提取内容"""
        import openpyxl
        
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text_parts = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # 添加工作表标题
            text_parts.append(f"【工作表: {sheet_name}】")
            
            if extract_mode == "legal_content":
                sheet_text = self._extract_legal_content_openpyxl(sheet)
            elif extract_mode == "law_articles":
                sheet_text = self._extract_law_articles_openpyxl(sheet)
            elif extract_mode == "table_structure":
                sheet_text = self._extract_table_structure_openpyxl(sheet)
            else:  # all_content
                sheet_text = self._extract_all_content_openpyxl(sheet)
            
            if sheet_text:
                text_parts.append(sheet_text)
        
        workbook.close()
        return '\n\n'.join(text_parts)
    
    def _extract_legal_content_openpyxl(self, sheet) -> str:
        """提取法律相关内容（优化版）"""
        content_parts = []
        
        # 法律文档关键词
        legal_keywords = ['条', '款', '项', '章', '节', '编', '篇', '规定', '办法', '条例', '法律', '法规']
        
        for row in sheet.iter_rows(values_only=True):
            if not any(row):  # 跳过空行
                continue
            
            row_text = []
            for cell_value in row:
                if cell_value is not None:
                    cell_str = str(cell_value).strip()
                    if cell_str:
                        row_text.append(cell_str)
            
            if row_text:
                row_content = " | ".join(row_text)
                
                # 检查是否包含法律关键词
                if any(keyword in row_content for keyword in legal_keywords):
                    content_parts.append(f"★ {row_content}")  # 标记重要内容
                else:
                    content_parts.append(row_content)
        
        return '\n'.join(content_parts)

    def _extract_law_articles_openpyxl(self, sheet) -> str:
        """提取法律条文结构（专门处理法规名称-条文内容格式）"""
        content_parts = []

        # 检查是否符合法律条文格式：A列法规名称，B列条文内容
        if sheet.max_row < 2 or sheet.max_column < 2:
            # 如果不符合格式，回退到普通法律内容提取
            return self._extract_legal_content_openpyxl(sheet)

        # 检查第一行是否是标题行
        first_row = [cell.value for cell in sheet[1]]
        is_header_row = any(
            str(cell).strip() in ['法规名称', '条文', '内容', '法律名称', '条款', '条文号', '第几条', '法规条文']
            for cell in first_row if cell is not None
        )

        start_row = 2 if is_header_row else 1

        # 收集所有数据来检测是否是特殊的法规-条文格式
        law_name = None
        articles = []
        same_law_name_count = 0

        for row_num in range(start_row, sheet.max_row + 1):
            row = sheet[row_num]

            # 获取A、B列的值
            current_law_name = str(row[0].value).strip() if row[0].value is not None else ""
            article_content = str(row[1].value).strip() if len(row) > 1 and row[1].value is not None else ""

            # 跳过空行或无效行
            if not article_content or article_content == "None":
                continue

            # 记录法规名称（取第一个非空的）
            if not law_name and current_law_name and current_law_name != "None":
                law_name = current_law_name

            # 统计相同法规名称的数量
            if current_law_name == law_name:
                same_law_name_count += 1

            # 从条文内容中提取条文号（如果存在）
            article_num = self._extract_article_number(article_content)

            articles.append({
                'law_name': current_law_name,
                'number': article_num,
                'content': article_content
            })

        # 检测是否是特殊格式：同一个法规名称重复出现，且条文内容包含条文号
        is_special_law_format = (
            law_name and
            same_law_name_count >= 2 and  # 至少有2行相同的法规名称
            len([a for a in articles if a['number']]) >= 2  # 至少有2个条文号
        )

        if is_special_law_format:
            # 特殊格式：法规名称作为第一个块，然后每个条文作为独立块
            content_parts.append(f"【LAW_NAME】")
            content_parts.append(law_name)
            content_parts.append("")

            # 然后添加每个条文作为独立的块
            for article in articles:
                if article['number']:
                    content_parts.append(f"【ARTICLE】{article['number']}")
                else:
                    content_parts.append(f"【ARTICLE】条文")
                content_parts.append(article['content'])
                content_parts.append("")
        else:
            # 普通格式：每行作为一个完整的条目
            for article in articles:
                if article['law_name'] and article['number']:
                    content_parts.append(f"【{article['law_name']} {article['number']}】")
                elif article['law_name']:
                    content_parts.append(f"【{article['law_name']}】")
                elif article['number']:
                    content_parts.append(f"【{article['number']}】")
                else:
                    content_parts.append(f"【条文】")

                content_parts.append(article['content'])
                content_parts.append("")

        return '\n'.join(content_parts)

    def _extract_article_number(self, content: str) -> str:
        """从条文内容中提取条文号"""
        import re

        # 匹配常见的条文号格式
        patterns = [
            r'^第[一二三四五六七八九十百千万\d]+条',  # 第X条
            r'^第[一二三四五六七八九十百千万\d]+款',  # 第X款
            r'^第[一二三四五六七八九十百千万\d]+项',  # 第X项
            r'^\d+\.',  # 数字编号
            r'^\(\d+\)',  # 括号数字
        ]

        for pattern in patterns:
            match = re.search(pattern, content)
            if match:
                return match.group(0)

        return ""

    def _extract_table_structure_openpyxl(self, sheet) -> str:
        """提取表格结构"""
        content_parts = []
        
        # 获取有数据的区域
        if sheet.max_row == 1 and sheet.max_column == 1:
            return ""
        
        # 处理表头
        header_row = []
        for cell in sheet[1]:
            if cell.value is not None:
                header_row.append(str(cell.value).strip())
            else:
                header_row.append("")
        
        if any(header_row):
            content_parts.append(f"表头: {' | '.join(header_row)}")
            content_parts.append("-" * 50)
        
        # 处理数据行
        for row_num in range(2, min(sheet.max_row + 1, 1000)):  # 限制行数避免过大
            row_data = []
            for cell in sheet[row_num]:
                if cell.value is not None:
                    row_data.append(str(cell.value).strip())
                else:
                    row_data.append("")
            
            if any(row_data):
                content_parts.append(" | ".join(row_data))
        
        return '\n'.join(content_parts)
    
    def _extract_all_content_openpyxl(self, sheet) -> str:
        """提取所有内容"""
        content_parts = []
        
        for row in sheet.iter_rows(values_only=True):
            if not any(row):
                continue
            
            row_text = []
            for cell_value in row:
                if cell_value is not None:
                    row_text.append(str(cell_value).strip())
            
            if row_text:
                content_parts.append(" | ".join(row_text))
        
        return '\n'.join(content_parts)
    
    def _extract_with_xlrd(self, file_path: str, extract_mode: str) -> str:
        """使用xlrd提取内容（用于.xls文件）"""
        import xlrd
        
        workbook = xlrd.open_workbook(file_path)
        text_parts = []
        
        for sheet_name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(sheet_name)
            
            text_parts.append(f"【工作表: {sheet_name}】")

            if extract_mode == "law_articles":
                sheet_text = self._extract_law_articles_xlrd(sheet)
            else:
                # 默认提取方式
                content_parts = []
                for row_idx in range(sheet.nrows):
                    row_data = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        if cell_value:
                            row_data.append(str(cell_value).strip())

                    if row_data:
                        content_parts.append(" | ".join(row_data))

                sheet_text = '\n'.join(content_parts) if content_parts else ""

            if sheet_text:
                text_parts.append(sheet_text)
        
        return '\n\n'.join(text_parts)

    def _extract_law_articles_xlrd(self, sheet) -> str:
        """使用xlrd提取法律条文结构"""
        content_parts = []

        # 检查是否符合法律条文格式：A列法规名称，B列条文内容
        if sheet.nrows < 2 or sheet.ncols < 2:
            # 如果不符合格式，回退到普通提取
            content_parts = []
            for row_idx in range(sheet.nrows):
                row_data = []
                for col_idx in range(sheet.ncols):
                    cell_value = sheet.cell_value(row_idx, col_idx)
                    if cell_value:
                        row_data.append(str(cell_value).strip())

                if row_data:
                    content_parts.append(" | ".join(row_data))

            return '\n'.join(content_parts)

        # 检查第一行是否是标题行
        first_row = [sheet.cell_value(0, col_idx) for col_idx in range(min(2, sheet.ncols))]
        is_header_row = any(
            str(cell).strip() in ['法规名称', '条文', '内容', '法律名称', '条款', '条文号', '第几条', '法规条文']
            for cell in first_row if cell
        )

        start_row = 1 if is_header_row else 0

        # 收集所有数据来检测是否是特殊的法规-条文格式
        law_name = None
        articles = []
        same_law_name_count = 0

        for row_idx in range(start_row, sheet.nrows):
            # 获取A、B列的值
            current_law_name = str(sheet.cell_value(row_idx, 0)).strip() if sheet.ncols > 0 else ""
            article_content = str(sheet.cell_value(row_idx, 1)).strip() if sheet.ncols > 1 else ""

            # 跳过空行或无效行
            if not article_content or article_content == "None" or not article_content.strip():
                continue

            # 记录法规名称（取第一个非空的）
            if not law_name and current_law_name and current_law_name != "None":
                law_name = current_law_name

            # 统计相同法规名称的数量
            if current_law_name == law_name:
                same_law_name_count += 1

            # 从条文内容中提取条文号（如果存在）
            article_num = self._extract_article_number(article_content)

            articles.append({
                'law_name': current_law_name,
                'number': article_num,
                'content': article_content
            })

        # 检测是否是特殊格式
        is_special_law_format = (
            law_name and
            same_law_name_count >= 2 and
            len([a for a in articles if a['number']]) >= 2
        )

        if is_special_law_format:
            # 特殊格式：法规名称作为第一个块，然后每个条文作为独立块
            content_parts.append(f"【LAW_NAME】")
            content_parts.append(law_name)
            content_parts.append("")

            for article in articles:
                if article['number']:
                    content_parts.append(f"【ARTICLE】{article['number']}")
                else:
                    content_parts.append(f"【ARTICLE】条文")
                content_parts.append(article['content'])
                content_parts.append("")
        else:
            # 普通格式
            for article in articles:
                if article['law_name'] and article['number']:
                    content_parts.append(f"【{article['law_name']} {article['number']}】")
                elif article['law_name']:
                    content_parts.append(f"【{article['law_name']}】")
                elif article['number']:
                    content_parts.append(f"【{article['number']}】")
                else:
                    content_parts.append(f"【条文】")

                content_parts.append(article['content'])
                content_parts.append("")

        return '\n'.join(content_parts)

    def _extract_with_pandas(self, file_path: str, extract_mode: str) -> str:
        """使用pandas提取内容（通用方法）"""
        import pandas as pd
        
        try:
            # 读取所有工作表
            excel_file = pd.ExcelFile(file_path)
            text_parts = []
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                text_parts.append(f"【工作表: {sheet_name}】")

                if extract_mode == "law_articles":
                    sheet_text = self._extract_law_articles_pandas(df)
                else:
                    # 默认转换为文本
                    content_parts = []
                    for _, row in df.iterrows():
                        row_data = []
                        for value in row:
                            if pd.notna(value):
                                row_data.append(str(value).strip())

                        if row_data:
                            content_parts.append(" | ".join(row_data))

                    sheet_text = '\n'.join(content_parts) if content_parts else ""

                if sheet_text:
                    text_parts.append(sheet_text)
            
            return '\n\n'.join(text_parts)
            
        except Exception as e:
            logger.error(f"pandas处理Excel失败: {e}")
            raise

    def _extract_law_articles_pandas(self, df) -> str:
        """使用pandas提取法律条文结构"""
        import pandas as pd

        content_parts = []

        # 检查是否符合法律条文格式：A列法规名称，B列条文内容
        if len(df.columns) < 2 or len(df) < 1:
            # 如果不符合格式，回退到普通提取
            content_parts = []
            for _, row in df.iterrows():
                row_data = []
                for value in row:
                    if pd.notna(value):
                        row_data.append(str(value).strip())

                if row_data:
                    content_parts.append(" | ".join(row_data))

            return '\n'.join(content_parts)

        # 检查第一行是否是标题行（通过列名或第一行数据判断）
        first_row_values = [str(val) for val in df.iloc[0].values if pd.notna(val)]
        is_header_row = any(
            val.strip() in ['法规名称', '条文', '内容', '法律名称', '条款', '条文号', '第几条', '法规条文']
            for val in first_row_values
        )

        # 如果第一行是标题，从第二行开始；否则从第一行开始
        start_idx = 1 if is_header_row else 0

        # 收集所有数据来检测是否是特殊的法规-条文格式
        law_name = None
        articles = []
        same_law_name_count = 0

        for idx in range(start_idx, len(df)):
            row = df.iloc[idx]

            # 获取A、B列的值
            current_law_name = str(row.iloc[0]).strip() if len(row) > 0 and pd.notna(row.iloc[0]) else ""
            article_content = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ""

            # 跳过空行或无效行
            if not article_content or article_content == "nan" or not article_content.strip():
                continue

            # 记录法规名称（取第一个非空的）
            if not law_name and current_law_name and current_law_name != "nan":
                law_name = current_law_name

            # 统计相同法规名称的数量
            if current_law_name == law_name:
                same_law_name_count += 1

            # 从条文内容中提取条文号（如果存在）
            article_num = self._extract_article_number(article_content)

            articles.append({
                'law_name': current_law_name,
                'number': article_num,
                'content': article_content
            })

        # 检测是否是特殊格式
        is_special_law_format = (
            law_name and
            same_law_name_count >= 2 and
            len([a for a in articles if a['number']]) >= 2
        )

        if is_special_law_format:
            # 特殊格式：法规名称作为第一个块，然后每个条文作为独立块
            content_parts.append(f"【LAW_NAME】")
            content_parts.append(law_name)
            content_parts.append("")

            for article in articles:
                if article['number']:
                    content_parts.append(f"【ARTICLE】{article['number']}")
                else:
                    content_parts.append(f"【ARTICLE】条文")
                content_parts.append(article['content'])
                content_parts.append("")
        else:
            # 普通格式
            for article in articles:
                if article['law_name'] and article['number']:
                    content_parts.append(f"【{article['law_name']} {article['number']}】")
                elif article['law_name']:
                    content_parts.append(f"【{article['law_name']}】")
                elif article['number']:
                    content_parts.append(f"【{article['number']}】")
                else:
                    content_parts.append(f"【条文】")

                content_parts.append(article['content'])
                content_parts.append("")

        return '\n'.join(content_parts)

    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """
        获取Excel文件信息
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            文件信息字典
        """
        info = {
            'file_path': file_path,
            'file_exists': os.path.exists(file_path),
            'file_size': 0,
            'file_type': Path(file_path).suffix.lower(),
            'is_excel': self.is_excel_file(file_path),
            'available_libraries': self.available_libraries,
            'recommended_library': None,
            'sheets_info': []
        }
        
        if info['file_exists']:
            info['file_size'] = os.path.getsize(file_path)
            
            # 推荐处理库
            if info['file_type'] in ['.xlsx', '.xlsm'] and self.available_libraries.get('openpyxl'):
                info['recommended_library'] = 'openpyxl'
            elif info['file_type'] == '.xls' and self.available_libraries.get('xlrd'):
                info['recommended_library'] = 'xlrd'
            elif self.available_libraries.get('pandas'):
                info['recommended_library'] = 'pandas'
            
            # 获取工作表信息
            if info['is_excel']:
                try:
                    info['sheets_info'] = self._get_sheets_info(file_path)
                except Exception as e:
                    logger.warning(f"获取工作表信息失败: {e}")
        
        return info
    
    def _get_sheets_info(self, file_path: str) -> List[Dict[str, Any]]:
        """获取工作表信息"""
        sheets_info = []
        
        try:
            if self.available_libraries.get('pandas'):
                import pandas as pd
                excel_file = pd.ExcelFile(file_path)
                
                for sheet_name in excel_file.sheet_names:
                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        sheets_info.append({
                            'name': sheet_name,
                            'rows': len(df),
                            'columns': len(df.columns),
                            'has_data': not df.empty
                        })
                    except Exception as e:
                        sheets_info.append({
                            'name': sheet_name,
                            'error': str(e)
                        })
        except Exception as e:
            logger.warning(f"获取工作表信息失败: {e}")
        
        return sheets_info
